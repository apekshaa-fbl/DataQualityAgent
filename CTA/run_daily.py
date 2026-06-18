"""
CTA Daily Automation — runs at 5pm IST (11:30 UTC) via cron.

Steps:
  1. Playwright exports Master Customers CSV from QuickSight
  2. Upload new dates to S3
  3. TOTP popup → connect to Snowflake
  4. Run CTA pipeline for today's date
  5. Query results + domains not in Firmable
  6. On Fridays: create Linear ticket with weekly missing domains
  7. Send Slack summary

Usage:
    python3 CTA/run_daily.py                   # full run (headless browser)
    python3 CTA/run_daily.py --skip-export     # skip Playwright, use latest CSV in Downloads
    python3 CTA/run_daily.py --date 2026-06-17 # process a specific date
    python3 CTA/run_daily.py --totp 123456     # skip TOTP popup
"""
import argparse
import csv
import io
import logging
import os
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))
from dotenv import load_dotenv
load_dotenv(override=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
)
logger = logging.getLogger("run_daily")

DOWNLOADS_DIR = Path.home() / "Downloads"
S3_BUCKET     = "fi-firmographics"
S3_PREFIX     = "data_quality/trial_audits/2026/"


# ── Snowflake connection (TOTP via macOS popup) ───────────────────────────────

def prompt_totp() -> str:
    """Show a native macOS dialog to collect the Snowflake TOTP code."""
    import subprocess
    script = (
        'display dialog "CTA Daily Pipeline\\n\\nEnter your Snowflake TOTP code:" '
        'default answer "" '
        'with title "Snowflake MFA" '
        'buttons {"Cancel", "OK"} default button "OK"'
    )
    result = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        raise RuntimeError("TOTP dialog cancelled or failed.")
    # output: "button returned:OK, text returned:123456"
    totp = result.stdout.strip().split("text returned:")[-1].strip()
    if not totp:
        raise RuntimeError("No TOTP entered.")
    return totp


def get_connection_with_totp(totp: str):
    """Connect to Snowflake using password + TOTP MFA."""
    from src.snowflake_conn import get_connection
    return get_connection(totp)


# ── S3 helpers ────────────────────────────────────────────────────────────────

def _s3_client():
    import boto3
    return boto3.client(
        "s3",
        aws_access_key_id=os.getenv("CTA_AWS_ACCESS_KEY_ID"),
        aws_secret_access_key=os.getenv("CTA_AWS_SECRET_ACCESS_KEY"),
    )


def get_existing_s3_keys() -> set:
    s3 = _s3_client()
    keys = set()
    paginator = s3.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=S3_BUCKET, Prefix=S3_PREFIX):
        for obj in page.get("Contents", []):
            keys.add(obj["Key"])
    return keys


def upload_date_to_s3(date_str: str, rows: list, fieldnames: list) -> str:
    s3 = _s3_client()
    _, mm, dd = date_str.split("-")
    filename = f"{mm}_{dd}_CTA.csv"
    s3_key   = f"{S3_PREFIX}{mm}/{filename}"

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    body = buf.getvalue().encode("utf-8")

    s3.put_object(Bucket=S3_BUCKET, Key=s3_key, Body=body, ContentType="text/csv")
    return s3_key


# ── CSV parsing ───────────────────────────────────────────────────────────────

def read_csv_by_date(csv_path: str) -> tuple[dict, list]:
    """Returns (groups_by_date, fieldnames)."""
    groups = defaultdict(list)
    fieldnames = None

    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(csv_path, newline="", encoding=enc) as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames
                for row in reader:
                    d = row.get("SUB_CREATED", "")[:10]
                    if d.startswith("2026"):
                        groups[d].append(row)
            break
        except UnicodeDecodeError:
            continue

    return dict(groups), fieldnames


# ── Snowflake queries ─────────────────────────────────────────────────────────

def get_monitoring_row(conn, run_date: str) -> dict | None:
    """Fetch the monitoring summary for a given date (all sub_statuses combined)."""
    from config import CTA_MONITORING_TABLE
    cur = conn.cursor()
    cur.execute(f"""
        SELECT
            SUM(total_customers)       AS total_customers,
            SUM(total_markets)         AS total_markets,
            SUM(in_firmable_count)     AS in_firmable,
            SUM(not_in_firmable_count) AS not_in_firmable,
            AVG(avg_field_coverage_pct)  AS field_coverage_pct,
            AVG(people_coverage_pct)     AS people_coverage_pct
        FROM {CTA_MONITORING_TABLE}
        WHERE run_date = %s
    """, (run_date,))
    row = cur.fetchone()
    cur.close()
    if not row:
        return None
    return {
        "total_customers":   int(row[0] or 0),
        "total_markets":     int(row[1] or 0),
        "in_firmable":       int(row[2] or 0),
        "not_in_firmable":   int(row[3] or 0),
        "field_coverage_pct":  round(row[4] or 0, 1),
        "people_coverage_pct": round(row[5] or 0, 1),
    }


def get_domains_not_in_firmable(conn, run_date: str) -> list[str]:
    """Return sorted list of customer domains not found in Firmable for a given date."""
    from config import CTA_STAGING_TABLE
    cur = conn.cursor()
    cur.execute(f"""
        SELECT DISTINCT customer_domain
        FROM {CTA_STAGING_TABLE}
        WHERE run_date = %s
          AND in_firmable = 'NOT_IN_FIRMABLE'
          AND customer_domain IS NOT NULL
          AND customer_domain != ''
        ORDER BY customer_domain
    """, (run_date,))
    domains = [row[0] for row in cur.fetchall()]
    cur.close()
    return domains


# ── Linear ────────────────────────────────────────────────────────────────────

LINEAR_PROJECT_ID = "edce723c-3f1e-49e8-8466-759a632e1291"
LINEAR_TEAM_ID    = "772eef07-c228-4ed5-aa8e-99c79877711b"
LINEAR_ASSIGNEE   = "185887b3-1659-4088-a418-7ee36b415908"  # apekshaa


def get_weekly_missing_domains(conn) -> list[dict]:
    """All distinct domains NOT in Firmable from the past 7 days, with first/last seen dates."""
    from config import CTA_STAGING_TABLE
    cur = conn.cursor()
    cur.execute(f"""
        SELECT
            LOWER(TRIM(customer_domain)) AS domain,
            MIN(run_date) AS first_seen,
            MAX(run_date) AS last_seen
        FROM {CTA_STAGING_TABLE}
        WHERE run_date >= DATEADD(day, -7, CURRENT_DATE())
          AND in_firmable = 'NOT_IN_FIRMABLE'
          AND customer_domain IS NOT NULL
          AND TRIM(customer_domain) != ''
        GROUP BY LOWER(TRIM(customer_domain))
        ORDER BY domain
    """)
    rows = [{"domain": r[0], "first_seen": str(r[1]), "last_seen": str(r[2])} for r in cur.fetchall()]
    cur.close()
    return rows


def create_linear_ticket(domains: list[dict], week_ending: str) -> str | None:
    """Create a Linear issue and attach an Excel file of missing domains. Returns issue URL."""
    import urllib.request, urllib.parse, json, tempfile
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    api_key = os.getenv("LINEAR_API_KEY")
    if not api_key:
        logger.warning("LINEAR_API_KEY not set — skipping Linear ticket creation")
        return None

    headers = {"Content-Type": "application/json", "Authorization": api_key}

    def gql(query, variables=None):
        body = json.dumps({"query": query, "variables": variables or {}}).encode()
        req = urllib.request.Request("https://api.linear.app/graphql", data=body, headers=headers)
        return json.loads(urllib.request.urlopen(req, timeout=15).read())

    # Step 1: Create the issue
    issue_data = gql("""
        mutation CreateIssue($input: IssueCreateInput!) {
          issueCreate(input: $input) {
            success
            issue { id url identifier }
          }
        }
    """, {"input": {
        "title": f"[CTA] Domains not in Firmable — Week ending {week_ending}",
        "description": (
            f"## Domains Not Found in Firmable\n\n"
            f"Week ending **{week_ending}** — {len(domains)} domain(s) could not be matched in Firmable.\n\n"
            f"See attached Excel file for the full list.\n\n"
            f"---\n*Auto-generated by the CTA Daily Pipeline*"
        ),
        "teamId": LINEAR_TEAM_ID,
        "projectId": LINEAR_PROJECT_ID,
        "assigneeId": LINEAR_ASSIGNEE,
        "priority": 3,
    }})
    issue = issue_data.get("data", {}).get("issueCreate", {}).get("issue", {})
    issue_id = issue.get("identifier")
    url = issue.get("url")
    if not issue_id:
        logger.error("Failed to create Linear issue")
        return None
    logger.info(f"Linear ticket created: {url}")

    # Step 2: Build Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Domains Not in Firmable"
    hdr_fill = PatternFill("solid", fgColor="0066CC")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(["Domain", "First Seen"], start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    from datetime import date as _date
    for i, d in enumerate(domains, start=2):
        ws.cell(row=i, column=1, value=d["domain"])
        cell = ws.cell(row=i, column=2, value=_date.fromisoformat(d["first_seen"]))
        cell.number_format = "YYYY-MM-DD"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        xlsx_path = tmp.name
    file_size = os.path.getsize(xlsx_path)
    filename = f"cta_domains_not_in_firmable_{week_ending}.xlsx"

    # Step 3: Get signed upload URL from Linear
    upload_data = gql("""
        mutation UploadFile($filename: String!, $contentType: String!, $size: Int!) {
          fileUpload(filename: $filename, contentType: $contentType, size: $size) {
            success
            uploadFile {
              uploadUrl
              assetUrl
              headers { key value }
            }
          }
        }
    """, {"filename": filename, "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "size": file_size})

    upload_file = upload_data.get("data", {}).get("fileUpload", {}).get("uploadFile", {})
    upload_url = upload_file.get("uploadUrl")
    asset_url  = upload_file.get("assetUrl")
    upload_headers = {h["key"]: h["value"] for h in upload_file.get("headers", [])}

    if not upload_url:
        logger.warning("Could not get Linear upload URL — skipping attachment")
        return url

    # Step 4: PUT the file
    with open(xlsx_path, "rb") as f:
        file_bytes = f.read()
    put_req = urllib.request.Request(upload_url, data=file_bytes, method="PUT")
    for k, v in upload_headers.items():
        put_req.add_header(k, v)
    urllib.request.urlopen(put_req, timeout=30)
    os.unlink(xlsx_path)

    # Step 5: Create attachment
    gql("""
        mutation CreateAttachment($input: AttachmentCreateInput!) {
          attachmentCreate(input: $input) { success }
        }
    """, {"input": {
        "issueId": issue.get("id"),
        "url": asset_url,
        "title": f"CTA Domains Not in Firmable — Week ending {week_ending}",
        "subtitle": f"{len(domains)} domains · Domain, First Seen",
    }})
    logger.info(f"Excel attachment uploaded to {issue_id}")
    return url


# ── Slack ─────────────────────────────────────────────────────────────────────

def send_slack_message(run_date: str, stats: dict, missing_domains: list[str], linear_url: str | None = None) -> None:
    import urllib.request
    import json

    webhook = os.getenv("SLACK_WEBHOOK_URL")
    if not webhook:
        logger.warning("SLACK_WEBHOOK_URL not set — skipping Slack notification")
        return

    # Format date as "17th June 2026"
    dt = datetime.strptime(run_date, "%Y-%m-%d")
    day = dt.day
    suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10 if day not in (11, 12, 13) else 0, "th")
    formatted_date = f"{day}{suffix} {dt.strftime('%B %Y')}"

    in_fbl_pct = round(stats['in_firmable'] / stats['total_customers'] * 100, 1) if stats['total_customers'] else 0

    footer = "Firmable Data Quality Bot  ·  Automated daily run"
    if linear_url:
        footer += f"  ·  <{linear_url}|Linear ticket>"

    payload = {
        "attachments": [
            {
                "color": "#0066CC",
                "fallback": f"CTA Daily — {formatted_date}: {stats['total_customers']} customers, {in_fbl_pct}% in Firmable",
                "mrkdwn_in": ["text", "footer"],
                "text": (
                    f"*:bar_chart: Customer Trial Audits  ·  {formatted_date}*\n"
                    f"*Customers:* {stats['total_customers']}    "
                    f"*Markets:* {stats['total_markets']}    "
                    f"*In Firmable:* {stats['in_firmable']} :white_check_mark:    "
                    f"*Not in Firmable:* {stats['not_in_firmable']} :x:\n"
                    f"*Coverage:* {in_fbl_pct}%    "
                    f"*Field Coverage:* {stats['field_coverage_pct']}%    "
                    f"*People Coverage:* {stats['people_coverage_pct']}%"
                ),
                "footer": footer,
            }
        ]
    }

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(webhook, data=data, headers={"Content-Type": "application/json"})
    urllib.request.urlopen(req, timeout=10)
    logger.info("Slack message sent.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="CTA Daily Automation")
    parser.add_argument("--skip-export", action="store_true", help="Skip Playwright, use latest CSV in Downloads")
    parser.add_argument("--date", default=None, help="Process this specific date only (YYYY-MM-DD)")
    parser.add_argument("--totp", default=None, help="Snowflake TOTP code (skips popup if provided)")
    args = parser.parse_args()

    target_date = args.date or date.today().isoformat()
    logger.info(f"=== CTA Daily Run — target date: {target_date} ===")

    # ── Step 1: Get CSV ────────────────────────────────────────────────────────
    if args.skip_export:
        csvs = sorted(DOWNLOADS_DIR.glob("Master_Customers_*.csv"), key=lambda f: f.stat().st_mtime)
        if not csvs:
            logger.error("No Master_Customers_*.csv found in Downloads. Exiting.")
            sys.exit(1)
        csv_path = str(csvs[-1])
        logger.info(f"Using existing CSV: {csv_path}")
    else:
        logger.info("Launching Playwright to export from QuickSight...")
        from playwright_export import export_csv
        csv_path = export_csv(headless=True)
        logger.info(f"Exported CSV: {csv_path}")

    # ── Step 2: Parse CSV ──────────────────────────────────────────────────────
    groups, fieldnames = read_csv_by_date(csv_path)
    if target_date not in groups:
        logger.warning(f"No rows for {target_date} in the CSV. Available dates: {sorted(groups.keys())[-5:]}")
        # Still continue — date may already be in S3 from a previous run
    else:
        logger.info(f"Found {len(groups[target_date])} rows for {target_date}")

    # ── Step 3: Upload today's date to S3 (if present in CSV) ─────────────────
    if target_date in groups:
        _, mm, dd = target_date.split("-")
        s3_key = f"{S3_PREFIX}{mm}/{mm}_{dd}_CTA.csv"
        existing = get_existing_s3_keys()
        if s3_key in existing:
            logger.info(f"S3 key already exists, re-uploading with latest data: {s3_key}")
        upload_date_to_s3(target_date, groups[target_date], fieldnames)
        logger.info(f"Uploaded to S3: {s3_key}")

    # ── Step 4: Connect to Snowflake (TOTP via popup) ─────────────────────────
    totp = args.totp or prompt_totp()
    logger.info("Connecting to Snowflake...")
    conn = get_connection_with_totp(totp)
    logger.info("Connected.")

    # ── Step 5: Run pipeline ───────────────────────────────────────────────────
    from src.cta_s3_loader import load_cta_file
    from src.cta_raw_table import write_cta_input
    from src.cta_staging import init_staging_table, run_cta_staging
    from src.cta_summary import init_summary_tables, run_cta_summary_by_status, run_cta_monitoring

    init_staging_table(conn)
    init_summary_tables(conn)

    _, mm, dd = target_date.split("-")
    s3_key = f"{S3_PREFIX}{mm}/{mm}_{dd}_CTA.csv"

    rows = load_cta_file(S3_BUCKET, s3_key)
    if not rows:
        logger.error(f"No rows loaded from S3 for {target_date}. Exiting.")
        conn.close()
        sys.exit(1)

    n = write_cta_input(conn, rows)
    conn.commit()
    logger.info(f"raw_input: {n} inserted")

    n_staged = run_cta_staging(conn, target_date)
    logger.info(f"staging: {n_staged} rows")

    run_cta_summary_by_status(conn, target_date)
    run_cta_monitoring(conn, target_date)
    logger.info("Summary + monitoring updated.")

    # ── Step 6: Query results ──────────────────────────────────────────────────
    stats = get_monitoring_row(conn, target_date)
    missing_domains = get_domains_not_in_firmable(conn, target_date)

    if not stats:
        conn.close()
        logger.error("No monitoring data found after pipeline run.")
        sys.exit(1)

    logger.info(f"Stats: {stats}")
    logger.info(f"Domains not in Firmable: {missing_domains}")

    # ── Step 7: Friday — create Linear ticket with weekly missing domains ──────
    linear_url = None
    if datetime.strptime(target_date, "%Y-%m-%d").weekday() == 4:  # Friday
        logger.info("Friday — fetching weekly missing domains for Linear ticket...")
        weekly_domains = get_weekly_missing_domains(conn)
        logger.info(f"Weekly missing domains: {len(weekly_domains)}")
        if weekly_domains:
            linear_url = create_linear_ticket(weekly_domains, target_date)

    conn.close()

    # ── Step 8: Send Slack ─────────────────────────────────────────────────────
    send_slack_message(target_date, stats, missing_domains, linear_url=linear_url)

    print("\n✅ Done.")
    print(f"   Date      : {target_date}")
    print(f"   Customers : {stats['total_customers']}")
    print(f"   In Fbl    : {stats['in_firmable']}")
    print(f"   Not in Fbl: {stats['not_in_firmable']}")
    if linear_url:
        print(f"   Linear    : {linear_url}")


if __name__ == "__main__":
    main()
